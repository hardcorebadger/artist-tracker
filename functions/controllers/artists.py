import hashlib
import json
import time
import copy
import traceback
import io
import uuid
from datetime import datetime

from firebase_admin import firestore
from google.cloud.firestore_v1 import FieldFilter
from sqlalchemy import select, and_, not_, or_, text, func
from sqlalchemy.orm import joinedload, contains_eager, aliased, subqueryload

from lib import Artist, get_user, LinkSource, ArtistLink, ArtistTag, OrganizationArtist, Evaluation, StatisticType, Statistic, \
    UserArtist, Attribution, pop_default


class ArtistController():

    def __init__(self, project_id, location):
        self.project_id = project_id
        self.location = location


    def queues(self, sql_session, app, uid):
        db = firestore.client(app)

        user_data = get_user(uid, db)
        count = self.build_query(uid, user_data, {}, db, sql_session, None, 'hide', True)
        return {
            "eval": count.filter(Artist.eval_queued_at != None).count(),
            "spotify": count.filter(Artist.spotify_queued_at != None).count(),
            "stats": count.filter(Artist.stats_queued_at != None).count(),

        }


    # def get_artists_test(self, data, app):
        # return (self.get_artists('9sRMdvFDUKVKckwpzeARiG6x2LG2', data, app, self.sql.get_session()))
    def get_artists(self, uid, data, app, sql_session, ids_only = False):
        id_lookup = data.get('id', None)

        try:
            # request schema from MUI
            # req.data = {'groupKeys': [], 'paginationModel': {'page': 0, 'pageSize': 10}, 'sortModel': [], 'filterModel': {'items': [], 'logicOperator': 'and', 'quickFilterValues': [], 'quickFilterLogicOperator': 'and'}, 'start': 0, 'end': 9}
            db = firestore.client(app)
            # How to get the user and the org IDs

            page = int(data.get('page', 0))
            page_size = int(data.get('pageSize', 20))

            # filters = data.get('filterModel', [])
            user_data = get_user(uid, db)
            # hashed_data = dict({'filters': filters, 'org': user_data.get('organization', None) })
            # json_string = json.dumps(hashed_data, sort_keys=True)

            # Create a hash object (using SHA-256 algorithm)
            # hash_object = hashlib.sha256(json_string.encode())

            # Get the hexadecimal representation of the hash
            # hex_digest = hash_object.hexdigest()
            # global count_by_query
            count = None
            # if count_by_query is None:
            #     count_by_query = dict()
            #
            # if hex_digest in count_by_query:
            #     count_object = count_by_query[hex_digest]
            #     if (time.time() - count_object['time']) < 360:
            #         count = count_object['count']
            if count is None and id_lookup is None and ids_only == False:
                count = self.build_query(uid, user_data, copy.deepcopy(dict(data)), db, sql_session, id_lookup, True).count()
                # count_by_query[hex_digest] = dict({"count": count, "time": time.time()})

            query = self.build_query(uid, user_data, copy.deepcopy(dict(data)), db, sql_session, id_lookup, False, ids_only)
            if not ids_only:
                query = query.limit(page_size).offset(page * page_size)

            artists_set = sql_session.scalars(query).unique()
            artists = None
            if id_lookup is not None:
                artists = list(map(lambda artist: artist.as_deep_dict(user_data.get('organization')), artists_set))
            elif ids_only == False:
                artists = list(map(lambda artist: artist.as_dict(user_data.get('organization')), artists_set))
            else:
                artists = list(artists_set)
                return artists
            db.close()
            if id_lookup is not None:
                return {
                    "artist": pop_default(artists, None),
                    "error": None
                }

            return {
                "rows": artists,
                "rowCount": count,
                "page": page,
                "pageSize": page_size,
                "filterModel": data.get('filterModel'),
                "sortModel": data.get('sortModel'),
                "error": None
            }
        except Exception as e:
            print (e)
            if id_lookup is not None:
                return {
                    "artist": None,
                    "error": traceback.format_exc()
                }
            return {
                "rows": [],
                "rowCount": 0,
                "page": data.get('page', 0),
                "pageSize": data.get('pageSize', 20),
                "filterModel": data.get('filterModel'),
                "sortModel": data.get('sortModel'),
                "error": traceback.format_exc()
            }

    def build_query(self, uid, user_data, data, db, sql_session, id_lookup, count = False, ids_only = False):
        query = sql_session.query(Artist)
        filter_model = data.get('filterModel')
        muted = 'hide'
        if filter_model is not None:
            muted = filter_model.get('muted')
        if not count and not ids_only:
            query = (select(Artist).options(
                joinedload(Artist.statistics).joinedload(Statistic.type, innerjoin=True).defer(StatisticType.created_at).defer(StatisticType.updated_at),
                joinedload(Artist.links, innerjoin=False).joinedload(ArtistLink.source, innerjoin=True).defer(LinkSource.logo),
                contains_eager(Artist.organizations),
                contains_eager(Artist.evaluation),
                joinedload(Artist.users, innerjoin=False),
                joinedload(Artist.tags, innerjoin=False),
            ))

        if id_lookup is not None and not ids_only:
            query = (select(Artist).options(
                joinedload(Artist.statistics).joinedload(Statistic.type, innerjoin=True).defer(
                    StatisticType.created_at).defer(StatisticType.updated_at),
                joinedload(Artist.links, innerjoin=False).joinedload(ArtistLink.source, innerjoin=True).defer(
                    LinkSource.logo),
                contains_eager(Artist.organizations),
                contains_eager(Artist.evaluation),
                joinedload(Artist.users, innerjoin=False),
                joinedload(Artist.tags, innerjoin=False),
                joinedload(Artist.attributions).joinedload(Attribution.playlist)
            ))
            query = query.filter(Artist.id == id_lookup)

        if ids_only:
            query = (select(Artist.id))

        org_filter = and_(OrganizationArtist.organization_id == user_data.get('organization'), OrganizationArtist.archived == False)
        if muted == 'hide' or muted is None:
            org_filter = and_(org_filter, OrganizationArtist.muted == False)
        elif muted == 'only':
            org_filter = and_(org_filter, OrganizationArtist.muted == True)

        sub_query = select(func.distinct(OrganizationArtist.artist_id)).filter(org_filter)
        query = query.filter(Artist.active == True)
        if not ids_only:
            query = query.outerjoin(Evaluation, Artist.evaluation)
            query = query.outerjoin(OrganizationArtist, and_(Artist.organizations, OrganizationArtist.organization_id == user_data.get('organization')))
        query = query.where(Artist.id.in_(sub_query))
        # query = query.filter(UserArtist.organization_id == user_data.get('organization'))
        # query = query.filter(or_(ArtistTag.organization_id == user_data.get('organization'), ArtistTag.organization_id == None))

        query, eval_dynamic, org_dynamic  = self.build_filters(user_data, data, query)

        # .where(Artist.organizations.any(OrganizationArtist.organization_id == user_data.get('organization'))))
        if id_lookup is None:
            query = self.build_sorts(data, query, count, user_data)

        return query

    def build_filters(self, user_data, data, query):
        filter_model = data.get('filterModel', False)
#         print(filter_model)
        eval_dynamic = None
        org_dynamic = None
        if not filter_model or len(filter_model) == 0:
            return query, eval_dynamic, org_dynamic

        filter_model = filter_model.get('items', list())
        for filter_field in filter_model:
            field = filter_field.get('field')
            operator = filter_field.get('operator')
            value = filter_field.get('value', None)
            if value is None and operator != 'isEmpty' and operator != 'isNotEmpty' and not field.startswith('tags.'):
                continue
            if (field.startswith('statistic.')):
                key_parts = field.split('.')
                statistic_id = key_parts[1].split('-')
                statistic_func = statistic_id[1]
                statistic_id = int(statistic_id[0])
                dynamic = aliased(Statistic)
                query = query.outerjoin(dynamic, and_(Artist.id == dynamic.artist_id,
                      dynamic.statistic_type_id == statistic_id))
                if value is not None:
                    if '.' in str(value):
                        value = float(value)
                    else:
                        value = int(value)
                if statistic_func in ['week_over_week', 'month_over_month']:
                    value /= 100
                query = self.build_condition(query, getattr(dynamic, statistic_func), operator, value)
            elif field == 'tags':
                tag_type = 1
                if field == 'tag_genre':
                    tag_type = 2

                if operator == 'is':
                    if value is None:
                        query = query.filter(~(Artist.tags.any(and_(ArtistTag.tag_type_id == tag_type, or_(ArtistTag.organization_id == user_data.get('organization'), ArtistTag.organization_id == None)))))
                    else:
                        query = query.filter(Artist.tags.any(and_(ArtistTag.tag == value, ArtistTag.tag_type_id == tag_type, or_(ArtistTag.organization_id == user_data.get('organization'), ArtistTag.organization_id == None))))
                elif operator == 'not':
                    if value is None:
                        query = query.filter(Artist.tags.any(and_(ArtistTag.tag_type_id == tag_type, or_(ArtistTag.organization_id == user_data.get('organization'), ArtistTag.organization_id == None))))
                    else:
                        query = query.filter(~(Artist.tags.any(and_(ArtistTag.tag == value, ArtistTag.tag_type_id == tag_type, or_(ArtistTag.organization_id == user_data.get('organization'), ArtistTag.organization_id == None)))))
                elif operator == 'isAnyOf' and value is not None and len(value) > 0:
                    query = query.filter(Artist.tags.any(and_(ArtistTag.tag.in_(value), ArtistTag.tag_type_id == tag_type, or_(ArtistTag.organization_id == user_data.get('organization'), ArtistTag.organization_id == None))))

            elif field.startswith('evaluation.'):
                eval_key = field.split('.')
                eval_key = eval_key[1]
                newValue = value

                eval_dynamic = aliased(Evaluation)
                column = getattr(eval_dynamic, eval_key)
                query = query.outerjoin(eval_dynamic, Artist.evaluation_id == eval_dynamic.id)
                query = self.build_condition(query, column, operator, newValue)
            elif field.startswith('organization.'):
                org_key = field.split('.')
                org_key = org_key[1]
                newValue = value
                if org_key == 'attribution_type':
                    org_key = 'last_playlist_id'
                    if newValue == 'true':
                        newValue = None
                        operator = 'isNotEmpty'
                    elif newValue == 'false':
                        newValue = None
                        operator = 'isEmpty'
                    else:
                        continue
                org_dynamic = aliased(OrganizationArtist)
                column = getattr(org_dynamic, org_key)
                query = query.outerjoin(org_dynamic, and_(Artist.id == org_dynamic.artist_id, org_dynamic.organization_id == user_data.get('organization'))).order_by(column)
                query = self.build_condition(query, column, operator, newValue)

            elif field == 'users':

                if operator == 'is':
                    query = query.filter(Artist.users.any(and_(UserArtist.user_id == value, UserArtist.organization_id == user_data.get('organization'))))
                elif operator == 'not':
                    query = query.filter(~Artist.users.any(and_(UserArtist.user_id == value, UserArtist.organization_id == user_data.get('organization'))))
                elif operator == 'isAnyOf' and len(value) > 0:
                    query = query.filter(Artist.users.any(and_(UserArtist.user_id.in_(value), UserArtist.organization_id == user_data.get('organization'))))

            else:
                column = Artist.__table__.columns[field]
                query = self.build_condition(query, column, operator, value)
        return query, eval_dynamic, org_dynamic

    def build_condition(self, query, column, operator, value):
        if value is None:
            if operator == '==' or operator == 'isEmpty':
                return query.filter(column == None)
            elif operator == '!=' or operator == 'isNotEmpty':
                return query.filter(column != None)
            else:
                return query
        if operator in ['after', 'before', 'onOrAfter', 'onOrBefore']:
            value = parse_datetime(value)
        if operator == 'isAnyOf':
            return query.filter(column.in_(value))
        elif operator == 'isNotOf':
            return query.filter(not_(column.in_(value)))
        elif operator == '>' or operator == 'after':
            return query.filter(column > value)
        elif operator == '<' or operator == 'before':
            return query.filter(column < value)
        elif operator == '>=' or operator == 'onOrAfter':
            return query.filter(column >= value)
        elif operator == '<=' or operator == 'onOrBefore':
            return query.filter(column <= value)
        elif operator == 'startsWith':
            search = '{}%'.format(escape_sql_search_text(value))
            return query.filter(column.like(search))
        elif operator == 'endsWith':
            search = '%{}'.format(escape_sql_search_text(value))
            return query.filter(column.like(search))
        elif operator == 'contains':
            return query.filter(column.contains(escape_sql_search_text(value)))
        elif operator == 'doesNotContain':
            return query.filter(not_(column.contains(value)))
        elif operator == '==' or operator == 'equals' or operator == 'is':
            return query.filter(column == value)
        elif operator == '!=' or operator == '<>' or operator == 'doesNotEqual' or operator == 'not':
            return query.filter(column != value)
        return query

    def build_sorts(self, data, query, count, user_data, eval_dynamic = None, org_dynamic = None):
        if count:
            return query
        if data.get('sortModel', False) and len(data['sortModel']) > 0:
            sorts = data['sortModel']
            for sort in sorts:
                sortFieldKey: str = sort['field']
                if (sortFieldKey.startswith('statistic.')):
                    statisticKeyParts = sortFieldKey.split('.')
                    statisticId = statisticKeyParts[1].split('-')
                    statisticFunc = statisticId[1]
                    statisticId = int(statisticId[0])
                    dynamic_sort = aliased(Statistic)
                    column = getattr(dynamic_sort, statisticFunc).asc()

                    if sort['sort'] == 'desc':
                        column = getattr(dynamic_sort, statisticFunc).desc()
                    query = query.outerjoin(dynamic_sort, Artist.id == dynamic_sort.artist_id).where(dynamic_sort.statistic_type_id == statisticId).order_by(column)
                    # query = query.filter(Artist.statistics.any(Statistic.statistic_type_id == statisticId & Statistic[statisticFunc] ))
                elif sortFieldKey.startswith('evaluation'):
                    evalKey = sortFieldKey.split('.')
                    evalKey = evalKey[1]

                    if eval_dynamic is None:
                        eval_dynamic = aliased(Evaluation)
                        query = query.outerjoin(eval_dynamic, Artist.evaluation_id == eval_dynamic.id)

                    column = getattr(eval_dynamic, evalKey).asc()

                    if sort['sort'] == 'desc':
                        column = getattr(eval_dynamic, evalKey).desc()
                    query = query.order_by(column)
                elif sortFieldKey.startswith('organization'):
                    orgKey = sortFieldKey.split('.')
                    orgKey = orgKey[1]
                    if org_dynamic is None:
                        org_dynamic = aliased(OrganizationArtist)
                        query = query.outerjoin(org_dynamic, and_(Artist.id == org_dynamic.artist_id, org_dynamic.organization_id == user_data.get('organization')))
                    column = getattr(org_dynamic, orgKey).asc()

                    if sort['sort'] == 'desc':
                        column = getattr(org_dynamic, orgKey).desc()
                    query = query.order_by(column)

                # elif sortFieldKey.startswith('user'):
                #     userKey = sortFieldKey.split('.')
                #     userKey = userKey[1]
                #     tableAliased = aliased(OrganizationArtist)
                #     column = getattr(tableAliased, userKey).asc()
                #
                #     if sort['sort'] == 'desc':
                #         column = getattr(tableAliased, userKey).desc()
                #     query = query.outerjoin(tableAliased, Artist.id == tableAliased.artist_id).where(tableAliased.organization_id == user_data['organization']).order_by(column)
                else:
                    column = Artist.__table__.columns[sortFieldKey].asc()

                    if sort['sort'] == 'desc':
                        column = Artist.__table__.columns[sortFieldKey].desc()
                    query = query.order_by(column)
        return query

    def get_label_type_counts(self, uid, data, app, sql_session):
        """Get counts of artists by label type (indie/major/unsigned) based on filter model."""
        try:
            db = firestore.client(app)
            user_data = get_user(uid, db)
            
            # Build base query with filters
            base_query = self.build_query(uid, user_data, copy.deepcopy(dict(data)), db, sql_session, id_lookup=None, count=True)
            
            # Create queries for each distributor type
            indie_query = base_query.filter(Evaluation.distributor_type == 1)
            major_query = base_query.filter(Evaluation.distributor_type == 2)
            diy_query = base_query.filter(Evaluation.distributor_type == 0)
            unknown_query = base_query.filter(or_(Evaluation.distributor_type == 3, Evaluation.id == None))
            
            return {
                "indie": indie_query.count(),
                "major": major_query.count(),
                "diy": diy_query.count(),
                "unknown": unknown_query.count()
            }
            
        except Exception as e:
            print(e)
            print(traceback.format_exc())
            return {
                "indie": 0,
                "major": 0,
                "diy": 0,
                "unknown": 0,
                "error": str(e)
            }

    def export_artists(self, sql_session, user_data, artists, column_order=None, export_format="csv"):
        """
        Exports artist data in either CSV or Excel format.
        
        Args:
            sql_session: SQL session
            user_data: User data dictionary
            artists: List of Artist objects to export
            column_order: Optional list of column names to include/order
            export_format: 'csv' or 'excel'
            
        Returns:
            Tuple of (file content, mimetype, filename)
        """
        try:
            # Get all statistic types and link sources for column headers
            stat_types_list = self._load_stat_types(sql_session)
            link_sources_list = self._load_link_sources(sql_session)
            
            # Create a set of all possible headers
            all_headers = {
                "artist_id", "name", "spotify_id", "avatar", "created_at", "updated_at", "onboarded",
                "distributor", "distributor_type", "label", "status", "back_catalog", 
                "evaluation_updated_at", "evaluation_created_at", "tags", "added_by", "added_on"
            }
            
            # Add link source headers to the set
            link_keys = {}
            for link_source in link_sources_list:
                key = f"link_{link_source.get('key')}"
                all_headers.add(key)
                link_keys[key] = link_source.get('key')
            
            # Add statistic headers to the set
            stat_keys = {}
            for stat_type in stat_types_list:
                source = stat_type.get('source')
                key = stat_type.get('key')
                stat_base_key = f"{source}_{key}"
                
                for stat_metric in ["latest", "previous", "week_over_week", "month_over_month", "min", "max", "avg"]:
                    stat_header = f"{stat_base_key}-{stat_metric}"
                    all_headers.add(stat_header)
                    if stat_header not in stat_keys:
                        stat_keys[stat_header] = {
                            "source": source,
                            "key": key,
                            "metric": stat_metric,
                            "full_key": stat_base_key,
                            "name": stat_type.get('name', '')
                        }
            
            # Create mapping for display names
            display_names = self._create_display_names(link_sources_list, stat_types_list)
            
            # If column_order is provided, filter and order headers accordingly
            if column_order and isinstance(column_order, list):
                # Map the dot notation column names to our internal format
                mapped_column_order = []
                for col in column_order:
                    # Handle evaluation fields
                    if col.startswith("evaluation."):
                        field = col.split(".")[1]
                        mapped_column_order.append(field)
                    # Handle organization fields
                    elif col.startswith("organization."):
                        field = col.split(".")[1]
                        if field == "created_at":
                            mapped_column_order.append("added_on")
                        else:
                            mapped_column_order.append(field)
                    # Handle users field
                    elif col == "users":
                        mapped_column_order.append("added_by")
                    # Handle tags field
                    elif col == "tags":
                        mapped_column_order.append("tags")
                    # Handle link fields
                    elif col.startswith("link_"):
                        mapped_column_order.append(col)
                    # Handle statistic fields with ID
                    elif col.startswith("statistic."):
                        parts = col.split(".")
                        parts_id = parts[1].split("-")
                        stat_type_id = str(parts_id[0])
                        metric = parts_id[1]
                        # Find the statistic type with this ID
                        for stat_type in stat_types_list:
                            if str(stat_type.get('id')) == stat_type_id:
                                source = stat_type.get('source')
                                key = stat_type.get('key')
                                mapped_column_order.append(f"{source}_{key}-{metric}")
                    else:
                        # For any other fields, keep as is
                        mapped_column_order.append(col.replace(".", "_"))
                
                column_order = ["artist_id", "name", "spotify_id"] + mapped_column_order
                # Only include headers that exist in all_headers
                headers = [header for header in column_order if header in all_headers]
            else:
                # Default headers if no column_order is provided
                headers = [
                    "artist_id", "name", "spotify_id", "avatar", "created_at", "updated_at", "onboarded",
                    "distributor", "distributor_type", "label", "status", "back_catalog", 
                    "evaluation_updated_at", "evaluation_created_at"
                ]
                
                # Add all link source headers
                for link_source in link_sources_list:
                    headers.append(f"link_{link_source.get('key')}")
                
                # Add all statistic headers
                for stat_type in stat_types_list:
                    source = stat_type.get('source')
                    key = stat_type.get('key')
                    headers.append(f"{source}_{key}-latest")
                    headers.append(f"{source}_{key}-previous")
                    headers.append(f"{source}_{key}-week_over_week")
                    headers.append(f"{source}_{key}-month_over_month")
                    headers.append(f"{source}_{key}-min")
                    headers.append(f"{source}_{key}-max")
                    headers.append(f"{source}_{key}-avg")
                
                # Add tags and added_by headers
                headers.append("tags")
                headers.append("added_by")
                headers.append("added_on")
            
            # Map for distributor type and status values
            distributor_type_map = {
                0: "DIY",
                1: "Indie",
                2: "Major",
                None: "Unknown"
            }
            
            status_map = {
                0: "Unsigned",
                1: "Signed",
                None: "Unknown"
            }
            
            back_catalog_map = {
                0: "Clean",
                1: "Dirty",
                None: "Unknown"
            }
            
            # Load user data for added_by field
            users_data = {}
            users = self._load_users(user_data.get('organization'))
            for user_info in users:
                users_data[user_info.get('id')] = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip()
            
            # Prepare data rows (field values for each artist)
            data_rows = []
            
            for artist in artists:
                # Create a dictionary to store all field values
                fields = {}
                
                # Basic artist fields
                fields["artist_id"] = artist.id
                fields["name"] = artist.name
                fields["spotify_id"] = artist.spotify_id
                fields["avatar"] = artist.avatar
                fields["created_at"] = artist.created_at.isoformat() if artist.created_at else ""
                fields["updated_at"] = artist.updated_at.isoformat() if artist.updated_at else ""
                fields["onboarded"] = artist.onboarded
                
                # Evaluation fields
                if artist.evaluation:
                    fields["distributor"] = artist.evaluation.distributor
                    fields["distributor_type"] = distributor_type_map.get(artist.evaluation.distributor_type, "Unknown")
                    fields["label"] = artist.evaluation.label
                    fields["status"] = status_map.get(artist.evaluation.status, "Unknown")
                    fields["back_catalog"] = back_catalog_map.get(artist.evaluation.back_catalog, "Unknown")
                    fields["evaluation_updated_at"] = artist.evaluation.updated_at.isoformat() if artist.evaluation.updated_at else ""
                    fields["evaluation_created_at"] = artist.evaluation.created_at.isoformat() if artist.evaluation.created_at else ""
                else:
                    fields["distributor"] = ""
                    fields["distributor_type"] = "Unknown"
                    fields["label"] = ""
                    fields["status"] = "Unknown"
                    fields["back_catalog"] = "Unknown"
                    fields["evaluation_updated_at"] = ""
                    fields["evaluation_created_at"] = ""
                
                # Get the needed link sources based on headers
                needed_link_keys = set()
                for header in headers:
                    if header.startswith("link_") and header in link_keys:
                        needed_link_keys.add(link_keys[header])
                
                # Only process links that are in the needed_link_keys set
                for link in artist.links:
                    if hasattr(link, 'source') and hasattr(link.source, 'key') and link.source.key in needed_link_keys:
                        fields[f"link_{link.source.key}"] = link.url
                
                # Fill in missing link fields with empty strings
                for header in headers:
                    if header.startswith("link_") and header not in fields:
                        fields[header] = ""
                
                # Get the needed statistic types based on headers
                needed_stat_keys = set()
                for header in headers:
                    if header in stat_keys:
                        needed_stat_keys.add(stat_keys[header]["full_key"])
                
                # Process statistics
                for stat in artist.statistics:
                    if hasattr(stat, 'type') and hasattr(stat.type, 'source') and hasattr(stat.type, 'key'):
                        key = f"{stat.type.source}_{stat.type.key}"
                        
                        # Only process stats that are in the needed_stat_keys set
                        if key in needed_stat_keys:
                            for metric in ["latest", "previous", "week_over_week", "month_over_month", "min", "max", "avg"]:
                                header = f"{key}-{metric}"
                                if header in headers:
                                    fields[header] = getattr(stat, metric)
                
                # Fill in missing stat fields with empty strings
                for header in headers:
                    if header in stat_keys and header not in fields:
                        fields[header] = ""
                
                # Tags field
                if "tags" in headers:
                    tags = [tag.tag for tag in artist.tags if tag.organization_id == user_data.get('organization')]
                    fields["tags"] = ",".join(tags) if tags else ""
                
                # Added by and added on fields
                org_artist = next((org for org in artist.organizations if org.organization_id == user_data.get('organization')), None)
                if org_artist:
                    if "added_by" in headers:
                        added_by_names = []
                        if hasattr(org_artist, 'added_by') and org_artist.added_by in users_data:
                            added_by_names.append(users_data[org_artist.added_by])
                        fields["added_by"] = ",".join(added_by_names) if added_by_names else ""
                    
                    if "added_on" in headers:
                        fields["added_on"] = org_artist.created_at.isoformat() if org_artist.created_at else ""
                else:
                    if "added_by" in headers:
                        fields["added_by"] = ""
                    if "added_on" in headers:
                        fields["added_on"] = ""
                
                data_rows.append(fields)
            
            # Now generate the output in the requested format
            if export_format.lower() == "excel":
                return self._generate_excel(headers, data_rows, stat_keys, display_names)
            else:
                return self._generate_csv(headers, data_rows, display_names)
                
        except Exception as e:
            print(str(e))
            print(traceback.format_exc())
            raise e
            
    def _create_display_names(self, link_sources_list, stat_types_list):
        """Create mapping of internal field names to display names"""
        display_names = {
            # Basic artist fields
            "artist_id": "Artist ID",
            "name": "Artist Name",
            "spotify_id": "Spotify ID",
            "avatar": "Avatar URL",
            "created_at": "Created At",
            "updated_at": "Updated At",
            "onboarded": "Onboarded",
            
            # Evaluation fields
            "distributor": "Distributor",
            "distributor_type": "Distributor Type",
            "label": "Label",
            "status": "Status",
            "back_catalog": "Back Catalog",
            "evaluation_updated_at": "Evaluation Updated At",
            "evaluation_created_at": "Evaluation Created At",
            
            # Organization fields
            "added_by": "Added By",
            "added_on": "Added On",
            "attribution_type": "Attribution Type",
            
            # Tags
            "tags": "Tags"
        }
        
        # Add link source display names
        for link_source in link_sources_list:
            key = f"link_{link_source.get('key')}"
            display_names[key] = f"{link_source.get('name', '').title()} Link"
        
        # Add statistic display names
        for stat_type in stat_types_list:
            source = stat_type.get('source')
            key = stat_type.get('key')
            name = stat_type.get('name', '')
            base_key = f"{source}_{key}"
            
            # Create display names for each metric
            display_names[f"{base_key}-latest"] = f"{name} (Latest)"
            display_names[f"{base_key}-previous"] = f"{name} (Previous)"
            display_names[f"{base_key}-week_over_week"] = f"{name} (Week/Week %)"
            display_names[f"{base_key}-month_over_month"] = f"{name} (Month/Month %)"
            display_names[f"{base_key}-min"] = f"{name} (Min)"
            display_names[f"{base_key}-max"] = f"{name} (Max)"
            display_names[f"{base_key}-avg"] = f"{name} (Avg)"
        
        return display_names
            
    def _generate_excel(self, headers, data_rows, stat_keys, display_names=None):
        """Generate Excel file from the data"""
        try:
            # Need to import here to avoid requiring openpyxl for non-Excel exports
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import uuid
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Artists"
            
            # Add headers with display names
            for col_idx, header in enumerate(headers, 1):
                # Use display name if available, otherwise use the original header
                header_display = display_names.get(header, header) if display_names else header
                cell = ws.cell(row=1, column=col_idx, value=header_display)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                
            # Add data rows
            for row_idx, fields in enumerate(data_rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = fields.get(header, "")
                    
                    # Convert UUID values to strings (especially for artist_id)
                    if isinstance(value, uuid.UUID):
                        value = str(value)
                    
                    # Check if this is a numeric field (statistics)
                    if header in stat_keys:
                        try:
                            if value not in (None, ""):
                                # For percentage fields
                                if stat_keys[header]["metric"] in ["week_over_week", "month_over_month"]:
                                    if isinstance(value, (int, float)):
                                        # Format as percentage
                                        ws.cell(row=row_idx, column=col_idx, value=value).number_format = "0.00%"
                                    else:
                                        ws.cell(row=row_idx, column=col_idx, value=value)
                                # For regular number fields
                                elif isinstance(value, (int, float)):
                                    ws.cell(row=row_idx, column=col_idx, value=value).number_format = "#,##0"
                                else:
                                    ws.cell(row=row_idx, column=col_idx, value=value)
                            else:
                                ws.cell(row=row_idx, column=col_idx, value="")
                        except Exception as e:
                            print(f"Error formatting cell: {str(e)}")
                            ws.cell(row=row_idx, column=col_idx, value=str(value))
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Adjust column widths based on content type
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                column_width = 0
                
                # Get header display name width
                header_display = display_names.get(header, header) if display_names else header
                header_width = len(str(header_display)) + 2  # Add padding
                column_width = max(column_width, header_width)
                
                # Set width based on content type
                if header == "artist_id":
                    # UUIDs are fixed length
                    column_width = max(column_width, 38)
                elif header in ["name", "distributor", "label"]:
                    # Names and labels get more space
                    column_width = max(column_width, 30)
                elif header == "avatar":
                    # URLs are longer
                    column_width = max(column_width, 40)
                elif header.startswith("link_"):
                    # URLs are longer
                    column_width = max(column_width, 40)
                elif header in stat_keys:
                    # Numbers are typically shorter
                    metric = stat_keys[header]["metric"]
                    if metric in ["week_over_week", "month_over_month"]:
                        # Percentages
                        column_width = max(column_width, 15)
                    else:
                        # Regular numbers
                        column_width = max(column_width, 12)
                elif header in ["created_at", "updated_at", "evaluation_created_at", "evaluation_updated_at", "added_on"]:
                    # Dates
                    column_width = max(column_width, 20)
                elif header == "tags":
                    # Tags can be long
                    column_width = max(column_width, 35)
                else:
                    # Default for other fields
                    column_width = max(column_width, 15)
                
                # Set the calculated width, with a reasonable max to avoid very wide columns
                ws.column_dimensions[column_letter].width = min(column_width, 50)
            
            # Freeze the header row
            ws.freeze_panes = "A2"
            
            # Save to a bytes buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return (output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'artist_export.xlsx')
        
        except Exception as e:
            print(f"Excel generation error: {str(e)}")
            print(traceback.format_exc())
            # Fall back to CSV if Excel generation fails
            return self._generate_csv(headers, data_rows, display_names)
    
    def _generate_csv(self, headers, data_rows, display_names=None):
        """Generate CSV file from the data"""
        try:
            # Helper function to escape CSV fields
            def escape_csv_field(field):
                if field is None:
                    return ""
                field_str = str(field)
                if "," in field_str or '"' in field_str or "\n" in field_str:
                    return f'"{field_str.replace('"', '""')}"'
                return field_str
            
            # Create CSV rows
            rows = []
            
            # Use display names for headers if available
            if display_names:
                header_row = [escape_csv_field(display_names.get(header, header)) for header in headers]
            else:
                header_row = [escape_csv_field(header) for header in headers]
                
            rows.append(",".join(header_row))
            
            for fields in data_rows:
                # Create the row based on the ordered headers
                row = [escape_csv_field(fields.get(header, "")) for header in headers]
                rows.append(",".join(row))
            
            # Create CSV content
            csv_content = "\n".join(rows)
            
            return (csv_content.encode('utf-8'), 'text/csv', 'artist_export.csv')
        
        except Exception as e:
            print(f"CSV generation error: {str(e)}")
            print(traceback.format_exc())
            raise e
    
    def _load_link_sources(self, sql_session):
        """Load all link sources"""
        sources = sql_session.query(LinkSource).all()
        return [{"id": source.id, "key": source.key, "name": source.display_name} for source in sources]
    
    def _load_stat_types(self, sql_session):
        """Load all statistic types"""
        stat_types = sql_session.query(StatisticType).all()
        return [{"id": stat_type.id, "source": stat_type.source, "key": stat_type.key, "name": stat_type.name} 
                for stat_type in stat_types]
    
    def _load_users(self, organization_id):
        """Load all users for an organization"""
        try:
            # Get all users in the organization
            user_records = []
            users_ref = firestore.client().collection('users')
            users = users_ref.where(filter=FieldFilter("organization", "==", organization_id)).get()
            
            for user in users:
                user_data = user.to_dict()
                user_data['id'] = user.id
                user_records.append(user_data)
                
            return user_records
        except Exception as e:
            print(f"Error loading users: {str(e)}")
            return []

def artist_joined_query():
    query = select(Artist).options(
        contains_eager(Artist.statistics),
        joinedload(Artist.links, innerjoin=False).joinedload(ArtistLink.source, innerjoin=True),
        joinedload(Artist.evaluation, innerjoin=False))
    # query = query.filter(Artist.active == True)
    return query


def artist_with_meta(sql_session, spotify_id = None, artist_id = None):
    query = select(Artist).options(
        joinedload(Artist.statistics, innerjoin=False).joinedload(Statistic.type, innerjoin=True),
        joinedload(Artist.links, innerjoin=False).joinedload(ArtistLink.source, innerjoin=True),
        joinedload(Artist.tags, innerjoin=False),
        joinedload(Artist.attributions_needing_notified, innerjoin=False)
    )

    if spotify_id is not None:
        query = query.where(Artist.spotify_id == spotify_id)
    if artist_id is not None:
        query = query.where(Artist.id == artist_id)
    return sql_session.scalars(query).first()


def parse_datetime(date_string):
    """Parses a datetime string with or without time."""
    try:
        # Try parsing with time
        return datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        try:
            # Try parsing without time
            return datetime.strptime(date_string, "%Y-%m-%d")
        except ValueError:
            raise ValueError("Invalid datetime format: {}".format(date_string))
def escape_sql_search_text(text: str) -> str:
    return text.replace("%", "\\%").replace("\\", "\\\\").replace("_", "\\_")